'''
Author: Chavez Cheong <https://github.com/ChavezCheong>
Purpose of Script: Automate downloading of PDF files from Ohio State Mortgage Enforcement Database using Selenium.
'''

# Handle Imports
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support import expected_conditions as EC
import time
from openpyxl import Workbook, load_workbook, workbook
import os.path
import glob

# Excel Constants <CHANGE THIS TO MATCH YOUR EXCEL FILE/CSV>
CASE_COL = 6
CRED_COL = 7
NAME_COL = 8
EA_COL = 10
FILENAME_COL = 12
COL_COUNT = 6951

# Web Page Constants <CHANGE THIS TO MATCH STATE>
STATE_URL = "https://apps2.com.ohio.gov/fiin/enforcementlookup/default.aspx"

# Directory Constants <CHANGE THIS TO MATCH YOUR SYSTEM>
FOLDER_PATH = r'C:\Users\chave\Desktop\Duke\APL2021_DataPlus\Data_Collection\OhioSelenium\Scraped Files'
FILE_TYPE = '\*'

# Helper Functions

def find_category(category_name):
    ''' Takes in the Entity Category Type and returns if it is to be searched as an individual/organization.

    Parameters:
    categoryName (string): Entity Category Type 

    Returns:
    string: "Organization" or "Individual"
    '''
    org_list = ["RM","SM","GL"]
    indiv_list = ["Loan Originator","Mortgage Loan Originator","OM"]
    if category_name in org_list:
        return "Org"
    elif category_name in indiv_list:
        return "Individual"     

def fill_in_org(name):
    '''Fill in organization name in entry field, if entity is an organization.'''
    org_button = driver.find_element_by_name("ctl00$MainContent$txtCompany")
    org_button.clear()
    org_button.send_keys(name)

def fill_in_name(name):
    '''Fill in individual name in entry field, if entity is an individual.'''
    last, first = name.split(", ")
    # Key in First Name
    first_name = driver.find_element_by_name("ctl00$MainContent$txtFirstName")
    first_name.clear()
    first_name.send_keys(first)
    # Key in Last Name
    last_name = driver.find_element_by_name("ctl00$MainContent$txtLastName")
    last_name.clear()
    last_name.send_keys(last)

def find_case_number(case_number):
    '''Returns case number from raw text.'''
    year, caseNo = caseNumber.split("-")
    return caseNo   

def return_custom_ea(ea):
    '''Handles Report and Recommendations in Ohio being saved as Division Orders.'''
    if ea == "REPORT AND RECOMMENDATION":
        return "DIVISION ORDER"
    else:
        return ea 

if __name__ == "__main__":

    # Load Excel
    workbook = load_workbook(filename=r"C:\Users\chave\Desktop\Duke\APL2021_DataPlus\Data_Collection\OhioSelenium\Mortgage Enforcement Actions Database (Trial).xlsx")
    sheet = workbook["Ohio"]

    # Test Variables
    # testList = [("JONES, WILLIAM", "Loan Originator")]
    # enforce = "NOTICE OF INTENT TO DENY RENEWAL"
    # enforce = "NOTICE OF INTENT TO DENY"
    # caseNo = "M2004-9991250"
    
    # Set up driver
    driver = webdriver.Chrome()
    driver.get(STATE_URL)
    input("Please please enter...")

    start = time.time()

    # Click the button to get into the page
    init_button = driver.find_element_by_name("ctl00$MainContent$btnContinue")
    init_button.click()

    # Loop through name and category for each row:
    for rows in range(1,COL_COUNT):
        print(rows)
        # Boolean check for file downloaded
        fileDownloaded = False
        
        # Set up variables
        caseNo = sheet.cell(row=rows, column=CASE_COL).value
        categoryName = sheet.cell(row=rows, column= CRED_COL).value
        name = sheet.cell(row=rows, column= NAME_COL).value
        enforce = sheet.cell(row=rows, column= EA_COL).value

        customEA = return_custom_ea(enforce)
                
        # print(name, categoryName)
        # Set up Page for Searches
        allButton = driver.find_element_by_name("ctl00$MainContent$rcbLicenseType$ctl00")
        allButton.click()

        # Set up Search
        category = find_category(categoryName)
        # print(category)
        if category == "Org":
            fill_in_org(name)
        if category == "Individual":
            fill_in_name(name)

        # Begin Search
        searchButton = driver.find_element_by_name("ctl00$MainContent$btnSearch")
        searchButton.click()

        # Open all available buttons
        buttonCount = 0
        buttonDefaultName1 = "ctl00$MainContent$dgResults$ctl00$ctl0"
        buttonDefaultName2 = "$GECBtnExpandColumn"
        try:
            while True:
                buttonName = buttonDefaultName1+str(4+3*buttonCount)+buttonDefaultName2
                expandButton = driver.find_element_by_name(buttonName)
                expandButton.click()
                buttonCount += 1
        except NoSuchElementException:
            if buttonCount == 0:
                driver.get(STATE_URL)
                continue
        
    
        # Find Stuff
        for fieldCount in range(1, buttonCount+1):
            rowCount = 0
            try:
                while True:
                    EADefaultName = f"//*[@id=\"ctl00_MainContent_dgResults_ctl00_ctl{str(3 * fieldCount + 3).zfill(2)}_Detail{str(fieldCount)}0__{str(fieldCount-1)}:0_{str(rowCount)}\"]/td[1]"
                    CaseDefaultName = f"//*[@id=\"ctl00_MainContent_dgResults_ctl00_ctl{str(3 * fieldCount + 3).zfill(2)}_Detail{str(fieldCount)}0__{str(fieldCount-1)}:0_{str(rowCount)}\"]/td[2]"
                    enforceAction = driver.find_element_by_xpath(EADefaultName).text
                    caseNumber = driver.find_element_by_xpath(CaseDefaultName).text
                    # Check if the case number, enforcement action type matches the Excel sheet and if we already downloaded the same file.
                    if enforceAction == customEA and find_case_number(caseNo) == caseNumber and not fileDownloaded:
                        # Get link to click
                        document = driver.find_element_by_xpath(CaseDefaultName+"/following-sibling::td[1]/*[1]")
                        # Handling different possible cases for Division Orders and Report and Recommendations
                        if enforce == "DIVISION ORDER" and customEA == "DIVISION ORDER" and "R&R" in document.text:
                            rowCount += 1
                            continue
                        if enforce == "DIVISION ORDER" and customEA == "DIVISION ORDER" and "RR" in document.text:
                            rowCount += 1
                            continue
                        if enforce == "REPORT AND RECOMMENDATION" and "R&R" in document.text:
                            document.click()
                            fileDownloaded = True
                            break
                        if enforce == "REPORT AND RECOMMENDATION" and "RR." in document.text:
                            document.click()
                            fileDownloaded = True
                            break
                        if enforce == "REPORT AND RECOMMENDATION" and not "R&R" in document.text:
                            rowCount += 1
                            continue
                        document.click()
                        fileDownloaded = True
                        break
                    else:
                        rowCount += 1
                        continue
            except NoSuchElementException:
                continue            
        
        # Fill in filename downloaded if true
        if fileDownloaded:
            try:
                time.sleep(2.5)
                files = glob.glob(FOLDER_PATH+FILE_TYPE)
                maxFile = max(files, key=os.path.getctime)
                fileName = os.path.splitext(os.path.basename(maxFile))[0]
                sheet.cell(row=rows, column=FILENAME_COL).value = fileName
            except FileNotFoundError:
                time.sleep(2.5)
                files = glob.glob(FOLDER_PATH+FILE_TYPE)
                maxFile = max(files, key=os.path.getctime)
                fileName = os.path.splitext(os.path.basename(maxFile))[0]
                sheet.cell(row=rows, column=FILENAME_COL).value = fileName
        driver.get(STATE_URL)
        
    # Save Excel Workbook
    workbook.save(r"C:\Users\chave\Desktop\Duke\APL2021_DataPlus\Data_Collection\OhioSelenium\test.xlsx")
    end = time.time()
    print(end - start)
        




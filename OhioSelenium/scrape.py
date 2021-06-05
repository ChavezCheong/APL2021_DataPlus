from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from openpyxl import Workbook, load_workbook, workbook
import os.path
import glob

# Excel Constants
CASE_COL = 6
CRED_COL = 7
NAME_COL = 8
EA_COL = 10
FILENAME_COL = 12

# Directory Constants
FOLDER_PATH = r'C:\Users\chave\Desktop\Duke\Data+\OhioSelenium\Scraped Files'
FILE_TYPE = '\*'

# Helper Functions

def getDownLoadedFileName(waitTime):
    driver.execute_script("window.open()")
    # switch to new tab
    driver.switch_to.window(driver.window_handles[-1])
    # navigate to chrome downloads
    driver.get('chrome://downloads')
    # define the endTime
    endTime = time.time()+waitTime
    while True:
        try:
            # get downloaded percentage
            downloadPercentage = driver.execute_script(
                "return document.querySelector('downloads-manager').shadowRoot.querySelector('#downloadsList downloads-item').shadowRoot.querySelector('#progress').value")
            # check if downloadPercentage is 100 (otherwise the script will keep waiting)
            if downloadPercentage == 100:
                # return the file name once the download is completed
                return driver.execute_script("return document.querySelector('downloads-manager').shadowRoot.querySelector('#downloadsList downloads-item').shadowRoot.querySelector('div#content  #file-link').text")
        except:
            pass
        time.sleep(1)
        if time.time() > endTime:
            break

def findCategory(categoryName):
    orgList = ["RM","SM","GL"]
    indivList = ["Loan Originator","Mortgage Loan Originator","OM"]
    if categoryName in orgList:
        return "Org"
    elif categoryName in indivList:
        return "Individual"     

def fillinOrg(name):
    # Key in Organization Name
    orgButton = driver.find_element_by_name("ctl00$MainContent$txtCompany")
    orgButton.clear()
    orgButton.send_keys(name)

def fillinName(name):
    # Format Name
    last, first = name.split(", ")
    # Key in First Name
    firstName = driver.find_element_by_name("ctl00$MainContent$txtFirstName")
    firstName.clear()
    firstName.send_keys(first)
    # Key in Last Name
    lastName = driver.find_element_by_name("ctl00$MainContent$txtLastName")
    lastName.clear()
    lastName.send_keys(last)

def findCaseNumber(caseNumber):
    year, caseNo = caseNumber.split("-")
    return caseNo   

def returnCustomEA(EA):
    if EA == "REPORT AND RECOMMENDATION":
        return "DIVISION ORDER"
    else:
        return EA 

if __name__ == "__main__":

    # Load Excel
    workbook = load_workbook(filename="Mortgage Enforcement Actions Database (Trial).xlsx")
    sheet = workbook["Ohio"]

    # Test Variables
    # testList = [("JONES, WILLIAM", "Loan Originator")]
    # enforce = "NOTICE OF INTENT TO DENY RENEWAL"
    # enforce = "NOTICE OF INTENT TO DENY"
    # caseNo = "M2004-9991250"
    

    # Set up driver
    driver = webdriver.Chrome()
    driver.get("https://apps2.com.ohio.gov/fiin/enforcementlookup/default.aspx")
    input("Please please enter...")

    start = time.time()

    # Click the button to get into the page
    initButton = driver.find_element_by_name("ctl00$MainContent$btnContinue")
    initButton.click()

    # for name, categoryName in testList:
    for rows in range(101,301):
        
        print(rows)
        # Boolean check for file downloaded
        fileDownloaded = False
        
        # Set up variables
        caseNo = sheet.cell(row=rows, column=CASE_COL).value
        categoryName = sheet.cell(row=rows, column= CRED_COL).value
        name = sheet.cell(row=rows, column= NAME_COL).value
        enforce = sheet.cell(row=rows, column= EA_COL).value

        customEA = returnCustomEA(enforce)
                
        # print(name, categoryName)
        # Set up Page for Searches
        allButton = driver.find_element_by_name("ctl00$MainContent$rcbLicenseType$ctl00")
        allButton.click()

        # Set up Search
        category = findCategory(categoryName)
        # print(category)
        if category == "Org":
            fillinOrg(name)
        if category == "Individual":
            fillinName(name)

        # Begin Search
        searchButton = driver.find_element_by_name("ctl00$MainContent$btnSearch")
        searchButton.click()

        # Open all available buttons
        buttonCount = 0
        buttonDefaultName1 = "ctl00$MainContent$dgResults$ctl00$ctl0"
        buttonDefaultName2 = "$GECBtnExpandColumn"
        try:
            while True:
                buttonName = buttonDefaultName1+str(4+3*buttonCount)+buttonDefaultName2
                expandButton = driver.find_element_by_name(buttonName)
                expandButton.click()
                buttonCount += 1
        except NoSuchElementException:
            if buttonCount == 0:
                driver.get("https://apps2.com.ohio.gov/fiin/enforcementlookup/default.aspx")
                continue
        
    
        # Find Stuff
        for fieldCount in range(1, buttonCount+1):
            rowCount = 0
            try:
                while True:
                    EADefaultName = f"//*[@id=\"ctl00_MainContent_dgResults_ctl00_ctl{str(3 * fieldCount + 3).zfill(2)}_Detail{str(fieldCount)}0__{str(fieldCount-1)}:0_{str(rowCount)}\"]/td[1]"
                    CaseDefaultName = f"//*[@id=\"ctl00_MainContent_dgResults_ctl00_ctl{str(3 * fieldCount + 3).zfill(2)}_Detail{str(fieldCount)}0__{str(fieldCount-1)}:0_{str(rowCount)}\"]/td[2]"
                    # print(EADefaultName)
                    # print(CaseDefaultName)
                    enforceAction = driver.find_element_by_xpath(EADefaultName).text
                    caseNumber = driver.find_element_by_xpath(CaseDefaultName).text
                    if enforceAction == customEA and findCaseNumber(caseNo) == caseNumber and not fileDownloaded:
                        document = driver.find_element_by_xpath(CaseDefaultName+"/following-sibling::td[1]/*[1]")
                        if enforce == "DIVISION ORDER" and customEA == "DIVISION ORDER" and "R&R" in document.text:
                            rowCount += 1
                            continue
                        if enforce == "DIVISION ORDER" and customEA == "DIVISION ORDER" and "RR" in document.text:
                            rowCount += 1
                            continue
                        if enforce == "REPORT AND RECOMMENDATION" and "R&R" in document.text:
                            document.click()
                            fileDownloaded = True
                            break
                        if enforce == "REPORT AND RECOMMENDATION" and "RR." in document.text:
                            document.click()
                            fileDownloaded = True
                            break
                        if enforce == "REPORT AND RECOMMENDATION" and not "R&R" in document.text:
                            rowCount += 1
                            continue
                        document.click()
                        fileDownloaded = True
                        break
                    else:
                        rowCount += 1
                        continue
            except NoSuchElementException:
                continue            
        
        # Fill in filename downloaded if true
        if fileDownloaded:
            time.sleep(2.5)
            files = glob.glob(FOLDER_PATH+FILE_TYPE)
            maxFile = max(files, key=os.path.getctime)
            fileName = os.path.splitext(os.path.basename(maxFile))[0]
            sheet.cell(row=rows, column=FILENAME_COL).value = fileName

        driver.get("https://apps2.com.ohio.gov/fiin/enforcementlookup/default.aspx")
        
    workbook.save("test.xlsx")
    end = time.time()
    print(end - start)
        



